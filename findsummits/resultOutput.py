import sys
import datetime
import os
import configparser
import csv
import openpyxl
import geojson
from operator import itemgetter
# 設定ファイル読み込み
config=configparser.ConfigParser()
config.read(f"{os.path.dirname(__file__)}/config.ini")
# エクセル作成
def makeXlsx():
    # peakColProminenceAllがあるか確認
    if os.path.isfile(config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"]):
        # あったらオープン
        with open(config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"]) as f:
            pcpa=[csvf for csvf in csv.reader(f)]
    else: # なければ終了
        print("not found:"+config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"])
        return
#
    wb=openpyxl.Workbook()  # 新規でワークブックを作成
    ws=wb.active    # ワークシートオブジェクトを取得
    # 見出し行を作成
#    print(ws["A1"].value)
#    print(ws.cell(1,1).value)
#    print(ws.cell(1,1).coordinate)
    ws["A1"].value="SummitCode"
    ws["B1"].value="SummitName"
    ws["C1"].value="AltM"
    ws["D1"].value="Longitude"
    ws["E1"].value="Latitude"
    ws["F1"].value="Peak AltM"
    ws["G1"].value="Peak Lon"
    ws["H1"].value="Peak Lat"
    ws["I1"].value="Col AltM"
    ws["J1"].value="Col Lon"
    ws["K1"].value="Col Lat"
    ws["L1"].value="Prominence"
    ws["M1"].value="Registration Date"
    ws["N1"].value="Update Date"
    # peakColProminenceAllの内容をシートに吐き出す
    pcpa.sort(key=itemgetter(12))   # SummitCodeで並べ替え
    for i,p in enumerate(pcpa,start=2):
        ws.cell(i,1).value=p[12]         # SummitCode
        ws.cell(i,2).value=p[13]         # SummitName
        ws.cell(i,3).value=int(p[14])    # AltM
        ws.cell(i,4).value=float(p[17])  # Longitude
        ws.cell(i,5).value=float(p[18])  # Latitude
        ws.cell(i,6).value=float(p[5])   # Peak AltM
        ws.cell(i,7).value=float(p[3])   # Peak Lon
        ws.cell(i,8).value=float(p[4])   # Peak Lat
        ws.cell(i,9).value=float(p[10])  # Col AltM
        ws.cell(i,10).value=float(p[8])  # Col Lon
        ws.cell(i,11).value=float(p[9])  # Col Lat
        ws.cell(i,12).value=f"=F{i}-I{i}" # Prominence 持ってるけどExcelで計算させる
        # Registration Date
        ws.cell(i,13).value=openpyxl.utils.datetime.from_ISO8601(p[19])
        # Update Date
        ws.cell(i,14).value=openpyxl.utils.datetime.from_ISO8601(p[20])
    # プロミネンス列に条件付き書式を設定
    yellowFill=openpyxl.styles.PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")
    ws.conditional_formatting.add(f"L2:L{len(pcpa)+1}",
        openpyxl.formatting.rule.CellIsRule(operator="lessThan",
            formula=[config["VAL"]["MINIMUM_PROMINENCE"]],
            stopIfTrue=True,fill=yellowFill)
    )
    # 終わったら保存
    wb.save(config["DIR"]["RESULT"]+"/"+config["FILE"]["SUMMITSPEAKCOLALL"])
#
    return
def makeGeojson():
    # peakColProminenceAllがあるか確認
    if os.path.isfile(config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"]):
        # あったらオープン
        with open(config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"]) as f:
            pcpa=[csvf for csvf in csv.reader(f)]
    else: # なければ終了
        print("not found:"+config["DIR"]["DATA"]+"/"+config["FILE"]["PEAKCOLPROMINENCEALL"])
        return
    # peakColProminenceAllの内容をgeoJSONに吐き出す
#    featureAll=[]
    featureCollectionAll=[]
    for p in pcpa:
        if float(p[14])>=1500:
            summitScore="10"
        elif float(p[14])>=1100:
            summitScore="8"
        elif float(p[14])>=850:
            summitScore="6"
        elif float(p[14])>=650:
            summitScore="4"
        elif float(p[14])>=500:
            summitScore="2"
        elif float(p[14])>=config["VAL"].getint("MINIMUM_PROMINENCE"):
            summitScore="1"
        else:
            assert False, "150mより低い山がある事は想定外。内容要確認"
        pointSummit=geojson.Point((float(p[17]),float(p[18])))
        featureSummit=geojson.Feature(
            geometry=pointSummit,id=p[12],
            properties={
                "name": p[13],
                "AltM": p[14]+"m",
				"_markerType": "Icon",
				"_iconUrl": config["URL"][f"ICONSUMMIT"+summitScore],
				"_iconSize": [16,16],
				"_iconAnchor": [8,8]
            }
        )
        if float(p[5])>=1500:
            peakScore="10"
        elif float(p[5])>=1100:
            peakScore="8"
        elif float(p[5])>=850:
            peakScore="6"
        elif float(p[5])>=650:
            peakScore="4"
        elif float(p[5])>=500:
            peakScore="2"
        elif float(p[5])>=config["VAL"].getint("MINIMUM_PROMINENCE"):
            peakScore="1"
        else:
            assert False, "150mより低い山がある事は想定外。内容要確認"
        pointPeak=geojson.Point((float(p[3]),float(p[4])))
        featurePeak=geojson.Feature(
            geometry=pointPeak,id=p[12]+"P",
            properties={
                "name": p[12]+" Peak in elevation tile",
                "AltM": p[5]+"m",
				"_markerType": "Icon",
				"_iconUrl": config["URL"][f"ICONPEAK"+peakScore],
				"_iconSize": [16,16],
				"_iconAnchor": [8,8]
            }
        )
        pointCol=geojson.Point((float(p[8]),float(p[9])))
        featureCol=geojson.Feature(
            geometry=pointCol,id=p[12]+"C",
            properties={
                "name": p[12]+" Col in elevation tile",
                "AltM": p[10]+"m",
				"_markerType": "Icon",
				"_iconUrl": config["URL"][f"ICONCOL"+peakScore],
				"_iconSize": [16,16],
				"_iconAnchor": [8,8]
            }
        )
        # 地理院地図でMultiLineStringが読み込めない様なので分割する
#        multiLineString=geojson.MultiLineString([
#            (float(p[17]),float(p[18])),(float(p[3]),float(p[4])),(float(p[8]),float(p[9]))
#        ])
#        featureMultiLine=geojson.Feature(geometry=multiLineString,id=p[12]+"ML")
        lineStringSp=geojson.LineString([
            (float(p[17]),float(p[18])),(float(p[3]),float(p[4]))
        ])
        featureLineSp=geojson.Feature(geometry=lineStringSp)
        lineStringPc=geojson.LineString([
            (float(p[3]),float(p[4])),(float(p[8]),float(p[9]))
        ])
        featureLinePc=geojson.Feature(geometry=lineStringPc)
        featureCollection=geojson.FeatureCollection(
#            [featureSummit,featurePeak,featureCol,featureMultiLine]
            [featureSummit,featurePeak,featureCol,lineStringSp,lineStringPc]
        )
#        print(pointSummit)
#        print(pointPeak)
#        print(pointCol)
#        print(lineStringSp)
#        print(lineStringPc)
#        print(featureSummit)
#        print(featurePeak)
#        print(featureCol)
#        print(featureLineSp)
#        print(featureLinePc)
#        print(featureCollection)
        featureCollectionAll.append(featureCollection)
    # 終わったらFeatureCollectionして保存
    with open(config["DIR"]["RESULT"]+"/"+config["FILE"]["GEOJSON"], 'w') as gjfile:
        # Output GeoJson file
        gjfile.write(geojson.dumps(featureCollectionAll,indent=2))
#
    return
#-------------Main---------------------------------
def main():
    print(f"{__name__}: Started @{datetime.datetime.now()}")
#
    # outputファイルを収めるディレクトリを先に作っておく
    os.makedirs(config["DIR"]["RESULT"] ,exist_ok=True)
#
    makeXlsx()
    makeGeojson()
#
    print(f"{__name__}: Finished @{datetime.datetime.now()}")
#---
if __name__ == '__main__':
    print(f"{sys.argv[0]}: Started @{datetime.datetime.now()}")
    args = sys.argv
    main()
    print(f"{sys.argv[0]}: Finished @{datetime.datetime.now()}")
